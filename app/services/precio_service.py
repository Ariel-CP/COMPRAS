from typing import Any, Dict, List, Optional
from datetime import date, datetime
import csv
import io
import logging
import re

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook, Workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from ..schemas.precio import PrecioImportResult


def _row_to_precio(row: Any) -> Dict[str, Any]:
    return {
        "id": row.id,
        "producto_id": row.producto_id,
        "producto_codigo": row.producto_codigo,
        "producto_nombre": row.producto_nombre,
        "proveedor_codigo": row.proveedor_codigo,
        "proveedor_nombre": row.proveedor_nombre,
        "fecha_precio": (
            row.fecha_precio.isoformat() if row.fecha_precio else None
        ),
        "precio_unitario": float(row.precio_unitario),
        "moneda": row.moneda,
        "origen": row.origen,
        "referencia_doc": row.referencia_doc,
        "notas": row.notas,
    }


def listar_precios_compra(
    db: Session,
    producto_id: Optional[int] = None,
    q: Optional[str] = None,
    proveedor: Optional[str] = None,
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
    limit: int = 50,
    offset: int = 0,
) -> List[Dict[str, Any]]:
    where = ["1=1"]
    params: Dict[str, Any] = {"limit": limit, "offset": offset}

    if producto_id is not None:
        where.append("h.producto_id = :pid")
        params["pid"] = producto_id

    if q:
        where.append(
            "(p.codigo LIKE :q OR p.nombre LIKE :q"
            " OR h.proveedor_codigo LIKE :q"
            " OR h.proveedor_nombre LIKE :q)"
        )
        params["q"] = f"%{q}%"

    if proveedor:
        where.append(
            "(h.proveedor_codigo LIKE :prov OR h.proveedor_nombre LIKE :prov)"
        )
        params["prov"] = f"%{proveedor}%"

    if desde is not None:
        where.append("h.fecha_precio >= :desde")
        params["desde"] = desde

    if hasta is not None:
        where.append("h.fecha_precio <= :hasta")
        params["hasta"] = hasta

    sql = text(
        """
        SELECT h.id, h.producto_id, h.proveedor_codigo, h.proveedor_nombre,
               h.fecha_precio, h.precio_unitario, h.moneda, h.origen,
               h.referencia_doc, h.notas,
               p.codigo AS producto_codigo, p.nombre AS producto_nombre
        FROM precio_compra_hist h
        JOIN producto p ON p.id = h.producto_id
        WHERE """
        + " AND ".join(where)
        + " ORDER BY h.fecha_precio DESC, h.id DESC"
        + " LIMIT :limit OFFSET :offset"
    )

    rows = db.execute(sql, params).fetchall()
    return [_row_to_precio(r) for r in rows]


ALLOWED_MONEDAS = {"ARS", "USD", "USD_MAY", "EUR"}
CURRENCY_ALIASES = {
    "ARS": "ARS",
    "PESO": "ARS",
    "PESOS": "ARS",
    "USD": "USD",
    "US": "USD",
    "DOLAR": "USD",
    "DOLARES": "USD",
    "EURO": "EUR",
    "EUROS": "EUR",
    "EUR": "EUR",
    "USD MAY": "USD_MAY",
    "USDMAY": "USD_MAY",
    "DOLAR MAY": "USD_MAY",
    "DOLARES MAY": "USD_MAY",
    "DOLAR MAYORISTA": "USD_MAY",
    "USD MAYORISTA": "USD_MAY",
}
ALLOWED_ORIGENES = {"ERP_FLEXXUS", "MANUAL", "OTRO"}
DEFAULT_PROVEEDOR_CODIGO = "PROV_GENERICO"
DEFAULT_PROVEEDOR_NOMBRE = "Proveedor Genérico"
DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%Y%m%d")
DATETIME_FORMATS = (
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y %H:%M",
    "%d/%m/%Y %I:%M:%S %p",
    "%d/%m/%Y %I:%M %p",
    "%Y-%m-%d %H:%M:%S",
)


def _normalize_datetime_text(raw: str) -> str:
    normalized = raw.replace("\xa0", " ").strip()
    normalized = " ".join(normalized.split())
    return re.sub(
        r"(?i)\b([ap])\.?\s*m\.?\b",
        lambda match: match.group(1).upper() + "M",
        normalized,
    )


def _sanitize_currency_key(raw: str) -> str:
    cleaned = raw.replace("\xa0", " ").strip().upper()
    cleaned = cleaned.replace("_", " ")
    cleaned = " ".join(cleaned.split())
    return re.sub(r"[^A-Z0-9 ]+", "", cleaned)


def _normalize_moneda_value(
    value: Any,
    default: Optional[str] = None
) -> Optional[str]:
    candidate = value if value not in (None, "") else default
    if candidate is None:
        return None
    key = _sanitize_currency_key(str(candidate))
    if not key:
        return None
    mapped = CURRENCY_ALIASES.get(key)
    return mapped if mapped in ALLOWED_MONEDAS else None


def _parse_fecha_precio(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    raw_text = str(value).strip()
    if not raw_text:
        return None
    cleaned = _normalize_datetime_text(raw_text)
    if not cleaned:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue
    for fmt in DATETIME_FORMATS:
        try:
            return datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue
    return None


def _decode_csv_content(content: bytes) -> str:
    if not content:
        raise HTTPException(status_code=400, detail="Archivo vacío")
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            return content.decode("latin-1")
        except UnicodeDecodeError as exc:
            raise HTTPException(
                status_code=400,
                detail=f"No se pudo decodificar el archivo: {exc}",
            ) from exc


def _normalize_row(row: Dict[str, Any]) -> Dict[str, Any]:
    normalized: Dict[str, Any] = {}
    for key, value in row.items():
        if key is None:
            continue
        norm_key = str(key).strip().lower()
        if isinstance(value, str):
            normalized[norm_key] = value.strip()
        else:
            normalized[norm_key] = value
    return normalized


def _parse_csv_rows(content: bytes) -> List[Dict[str, Any]]:
    decoded = _decode_csv_content(content)
    text_stream = io.StringIO(decoded)
    sample = text_stream.read(2048)
    text_stream.seek(0)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error as exc:
        raise HTTPException(
            status_code=400,
            detail=f"No se pudo detectar el formato del CSV: {exc}",
        ) from exc
    reader = csv.DictReader(text_stream, dialect=dialect)
    return [_normalize_row(row) for row in reader]


def _parse_xlsx_rows(content: bytes) -> List[Dict[str, Any]]:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Archivo XLSX inválido: {exc}",
        ) from exc
    ws = wb.active
    try:
        headers = [
            str(c.value).strip().lower() if c.value is not None else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]
    except StopIteration:
        return []
    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row_dict: Dict[str, Any] = {}
        for idx, cell in enumerate(r):
            key = headers[idx] if idx < len(headers) else f"col_{idx}"
            if isinstance(cell, str):
                row_dict[key] = cell.strip()
            else:
                row_dict[key] = cell
        rows.append(row_dict)
    return rows


def _get_producto_id(
    db: Session, cache: Dict[str, Optional[int]], codigo: str
) -> Optional[int]:
    if codigo in cache:
        return cache[codigo]
    res = db.execute(
        text("SELECT id FROM producto WHERE codigo = :codigo"),
        {"codigo": codigo},
    ).first()
    cache[codigo] = int(res[0]) if res else None
    return cache[codigo]


def generar_template_precios() -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "precios"
    ws.append(
        [
            "producto_codigo",
            "proveedor_codigo",
            "proveedor_nombre",
            "fecha_precio",
            "precio_unitario",
            "moneda",
            "origen",
            "referencia_doc",
            "notas",
        ]
    )
    ws.append(
        [
            "MAT-001",
            "PROV01",
            "Proveedor Demo",
            datetime.today().date().isoformat(),
            "123.45",
            "ARS",
            "MANUAL",
            "OC-123",
            "Observaciones",
        ]
    )
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def importar_precios_desde_archivo(
    db: Session, archivo: UploadFile
) -> PrecioImportResult:
    filename = archivo.filename or ""
    content = archivo.file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Archivo vacío")

    if filename.lower().endswith(".xlsx"):
        rows = _parse_xlsx_rows(content)
    elif filename.lower().endswith(".csv") or filename == "":
        rows = _parse_csv_rows(content)
    else:
        raise HTTPException(
            status_code=400,
            detail="Extensión no soportada (usar .csv o .xlsx)",
        )

    if not rows:
        return PrecioImportResult(
            insertados=0,
            actualizados=0,
            rechazados=0,
            errores=["Archivo vacío"],
        )

    required = {
        "producto_codigo",
        "fecha_precio",
        "precio_unitario",
        "moneda",
    }
    missing = required - set(rows[0].keys())
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Faltan columnas requeridas: {', '.join(sorted(missing))}",
        )

    insertados = actualizados = rechazados = 0
    errores: List[str] = []
    cache: Dict[str, Optional[int]] = {}

    try:
        for idx, raw_row in enumerate(rows, start=2):
            row = _normalize_row(raw_row)
            codigo = row.get("producto_codigo", "") or ""
            prov_codigo = (
                row.get("proveedor_codigo", "") or DEFAULT_PROVEEDOR_CODIGO
            )
            if not codigo:
                rechazados += 1
                errores.append(f"Fila {idx}: producto_codigo vacío")
                continue

            prod_id = _get_producto_id(db, cache, codigo)
            if not prod_id:
                rechazados += 1
                errores.append(
                    f"Fila {idx}: producto no encontrado ({codigo})"
                )
                continue

            fecha_precio = _parse_fecha_precio(row.get("fecha_precio"))
            if not fecha_precio:
                rechazados += 1
                errores.append(f"Fila {idx}: fecha_precio inválida")
                continue

            precio_raw = row.get("precio_unitario")
            try:
                precio = float(str(precio_raw).replace(",", "."))
                if precio <= 0:
                    raise ValueError()
            except (ValueError, TypeError):
                rechazados += 1
                errores.append(f"Fila {idx}: precio_unitario inválido")
                continue

            moneda_raw = row.get("moneda")
            moneda = _normalize_moneda_value(moneda_raw, "ARS")
            if not moneda:
                rechazados += 1
                errores.append(
                    f"Fila {idx}: moneda inválida ({moneda_raw or ''})"
                )
                continue

            origen = (row.get("origen") or "MANUAL").upper()
            if origen not in ALLOWED_ORIGENES:
                rechazados += 1
                errores.append(f"Fila {idx}: origen inválido ({origen})")
                continue

            prov_nombre = (
                row.get("proveedor_nombre") or DEFAULT_PROVEEDOR_NOMBRE
            )
            referencia = row.get("referencia_doc") or None
            notas = row.get("notas") or None

            existing = db.execute(
                text(
                    "SELECT id FROM precio_compra_hist "
                    "WHERE producto_id=:pid AND proveedor_codigo=:prov "
                    "AND fecha_precio=:fecha AND moneda=:moneda"
                ),
                {
                    "pid": prod_id,
                    "prov": prov_codigo,
                    "fecha": fecha_precio,
                    "moneda": moneda,
                },
            ).first()

            if existing:
                db.execute(
                    text(
                        "UPDATE precio_compra_hist SET "
                        "precio_unitario=:precio, proveedor_nombre=:prov_nom, "
                        "origen=:origen, referencia_doc=:ref, notas=:notas "
                        "WHERE id=:id"
                    ),
                    {
                        "precio": precio,
                        "prov_nom": prov_nombre,
                        "origen": origen,
                        "ref": referencia,
                        "notas": notas,
                        "id": existing[0],
                    },
                )
                actualizados += 1
            else:
                db.execute(
                    text(
                        "INSERT INTO precio_compra_hist "
                        "(producto_id, proveedor_codigo, proveedor_nombre, "
                        "fecha_precio, precio_unitario, moneda, origen, "
                        "referencia_doc, notas) VALUES "
                        "(:pid, :prov, :prov_nom, :fecha, :precio, :moneda, "
                        ":origen, :ref, :notas)"
                    ),
                    {
                        "pid": prod_id,
                        "prov": prov_codigo,
                        "prov_nom": prov_nombre,
                        "fecha": fecha_precio,
                        "precio": precio,
                        "moneda": moneda,
                        "origen": origen,
                        "ref": referencia,
                        "notas": notas,
                    },
                )
                insertados += 1

        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        logging.exception("Error importando precios")
        raise HTTPException(
            status_code=500,
            detail=f"Error importando precios: {exc}",
        ) from exc

    return PrecioImportResult(
        insertados=insertados,
        actualizados=actualizados,
        rechazados=rechazados,
        errores=errores,
    )
