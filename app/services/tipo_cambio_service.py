from datetime import date, datetime
from typing import Any, List, Optional, Sequence, Tuple, Union
from sqlalchemy import text
from sqlalchemy.engine import Connection
from sqlalchemy.orm import Session

from app.schemas.tipo_cambio import (
    TipoCambioCreate,
    TipoCambioUpdate,
    TipoCambioFiltro,
)


SQLConn = Union[Connection, Session]


def listar_tipos_cambio(
    conn: SQLConn, filtro: TipoCambioFiltro
) -> List[dict]:
    """Lista tipos de cambio con filtros opcionales."""
    clauses = []
    params: dict[str, object] = {}
    if filtro.moneda:
        clauses.append("moneda = :moneda")
        params["moneda"] = filtro.moneda
    if filtro.tipo:
        clauses.append("tipo = :tipo")
        params["tipo"] = filtro.tipo
    if filtro.desde:
        clauses.append("fecha >= :desde")
        params["desde"] = filtro.desde
    if filtro.hasta:
        clauses.append("fecha <= :hasta")
        params["hasta"] = filtro.hasta

    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    sql = f"""
    SELECT id, fecha, moneda, tipo, tasa, origen, notas, fecha_creacion
    FROM tipo_cambio_hist
    {where_sql}
    ORDER BY fecha DESC, moneda, tipo
    """.strip()
    result = conn.execute(text(sql), params).mappings()
    rows = []
    for r in result:
        row_dict = dict(r)
        # Convertir tipos no serializables a JSON
        if row_dict.get("fecha"):
            row_dict["fecha"] = row_dict["fecha"].isoformat()
        if row_dict.get("fecha_creacion"):
            row_dict["fecha_creacion"] = row_dict["fecha_creacion"].isoformat()
        if row_dict.get("tasa"):
            row_dict["tasa"] = float(row_dict["tasa"])
        rows.append(row_dict)
    return rows


def obtener_resumen_ultimas_tasas(conn: SQLConn) -> List[dict]:
    """Devuelve la última fecha y tasa por moneda/tipo."""
    sql = text(
        """
        SELECT t.moneda, t.tipo, t.fecha, t.tasa
        FROM tipo_cambio_hist t
        INNER JOIN (
            SELECT moneda, tipo, MAX(fecha) AS max_fecha
            FROM tipo_cambio_hist
            GROUP BY moneda, tipo
        ) ult
            ON ult.moneda = t.moneda
            AND ult.tipo = t.tipo
            AND ult.max_fecha = t.fecha
        ORDER BY t.moneda, t.tipo
        """
    )
    result = conn.execute(sql).mappings()
    resumen: List[dict] = []
    for row in result:
        data = dict(row)
        data["fecha"] = data["fecha"].isoformat()
        data["tasa"] = float(data["tasa"])
        resumen.append(data)
    return resumen


def upsert_tipo_cambio(
    conn: SQLConn, data: TipoCambioCreate
) -> Tuple[bool, int]:
    """Inserta o actualiza (si existe por PK única) un tipo de cambio.
    Devuelve (insertado, id).
    """
    # Intentar obtener existente
    sel = text("""
        SELECT id FROM tipo_cambio_hist
        WHERE moneda=:moneda AND fecha=:fecha AND tipo=:tipo
    """)
    existing = conn.execute(
        sel,
        {"moneda": data.moneda, "fecha": data.fecha, "tipo": data.tipo},
    ).fetchone()
    if existing:
        upd = text("""
            UPDATE tipo_cambio_hist
            SET tasa=:tasa, origen=:origen, notas=:notas
            WHERE id=:id
        """)
        conn.execute(
            upd,
            {
                "tasa": data.tasa,
                "origen": data.origen,
                "notas": data.notas,
                "id": existing.id,
            },
        )
        return False, existing.id
    ins = text("""
        INSERT INTO tipo_cambio_hist (fecha, moneda, tipo, tasa, origen, notas)
        VALUES (:fecha, :moneda, :tipo, :tasa, :origen, :notas)
    """)
    res = conn.execute(ins, {
        "fecha": data.fecha,
        "moneda": data.moneda,
        "tipo": data.tipo,
        "tasa": data.tasa,
        "origen": data.origen,
        "notas": data.notas,
    })
    return True, res.lastrowid  # type: ignore[attr-defined]


def actualizar_tipo_cambio(
    conn: SQLConn, id_: int, data: TipoCambioUpdate
) -> bool:
    set_parts = []
    params: dict[str, Any] = {"id": id_}
    if data.tasa is not None:
        set_parts.append("tasa=:tasa")
        params["tasa"] = data.tasa
    if data.origen is not None:
        set_parts.append("origen=:origen")
        params["origen"] = data.origen
    if data.notas is not None:
        set_parts.append("notas=:notas")
        params["notas"] = data.notas
    if not set_parts:
        return False
    sql = text(
        f"UPDATE tipo_cambio_hist SET {', '.join(set_parts)} WHERE id=:id"
    )
    res = conn.execute(sql, params)
    rowcount = getattr(res, "rowcount", 0)
    return bool(rowcount)


def obtener_por_id(conn: SQLConn, id_: int) -> Optional[dict]:
    sql = text("""
        SELECT id, fecha, moneda, tipo, tasa, origen, notas, fecha_creacion
        FROM tipo_cambio_hist WHERE id=:id
    """)
    row = conn.execute(sql, {"id": id_}).mappings().first()
    if not row:
        return None
    row_dict = dict(row)
    # Convertir tipos no serializables a JSON
    if row_dict.get("fecha"):
        row_dict["fecha"] = row_dict["fecha"].isoformat()
    if row_dict.get("fecha_creacion"):
        row_dict["fecha_creacion"] = row_dict["fecha_creacion"].isoformat()
    if row_dict.get("tasa"):
        row_dict["tasa"] = float(row_dict["tasa"])
    return row_dict


def bulk_import_csv(
    conn: SQLConn,
    contenido_csv: str,
    moneda: str = "USD",
    tipo: str = "VENTA",
    origen: str = "MANUAL",
) -> Tuple[int, int, List[str]]:
    """Importa CSV en memoria.
    Formato esperado (sin encabezado o con encabezado reconocible):
    fecha,tasa
    2025-01-02,1234.56
    Se ignoran líneas vacías. Se aceptan separadores "," o ";".
    """
    import csv
    from io import StringIO

    insertados = 0
    actualizados = 0
    errores: List[str] = []

    # Normalizar separadores ; a ,
    normalized = contenido_csv.replace(";", ",")
    f = StringIO(normalized)
    reader = csv.reader(f)

    # Detectar encabezado
    try:
        first_row_peek = next(reader)
    except StopIteration:
        return 0, 0, ["Archivo vacío"]

    has_header = bool(
        first_row_peek and any(h.lower() in ("fecha", "tasa") for h in first_row_peek)
    )

    if has_header:
        # Re-crear reader para después del encabezado
        f.seek(0)
        dict_reader = csv.DictReader(f)
        for idx, row in enumerate(dict_reader, start=2):
            try:
                fecha_str = row.get("fecha") or row.get("Fecha")
                tasa_str = row.get("tasa") or row.get("Tasa")
                if not fecha_str or not tasa_str:
                    errores.append(f"Línea {idx}: faltan campos")
                    continue
                fecha_dt = date.fromisoformat(fecha_str.strip())
                tasa_val = float(tasa_str.strip())
                creado, _ = upsert_tipo_cambio(
                    conn,
                    TipoCambioCreate(
                        fecha=fecha_dt,
                        moneda=moneda,
                        tipo=tipo,
                        tasa=tasa_val,
                        origen=origen,
                        notas=None,
                    ),
                )
                if creado:
                    insertados += 1
                else:
                    actualizados += 1
            except Exception as e:  # noqa: BLE001
                errores.append(f"Línea {idx}: {e}")
    else:
        # Sin encabezado: columnas esperadas fecha,tasa
        f.seek(0)
        for idx, cols in enumerate(csv.reader(f), start=1):
            if not cols or len(cols) < 2:
                continue
            try:
                fecha_dt = date.fromisoformat(cols[0].strip())
                tasa_val = float(cols[1].strip())
                creado, _ = upsert_tipo_cambio(
                    conn,
                    TipoCambioCreate(
                        fecha=fecha_dt,
                        moneda=moneda,
                        tipo=tipo,
                        tasa=tasa_val,
                        origen=origen,
                        notas=None,
                    ),
                )
                if creado:
                    insertados += 1
                else:
                    actualizados += 1
            except Exception as e:  # noqa: BLE001
                errores.append(f"Línea {idx}: {e}")

    return insertados, actualizados, errores


def bulk_import_xlsx(
    conn: SQLConn,
    file_bytes: bytes,
    moneda: str = "USD",
    tipo: str = "VENTA",
    origen: str = "MANUAL",
    sheet_name: str | None = None,
) -> Tuple[int, int, List[str]]:
    """Importa datos desde un archivo XLSX en bytes.
    Se espera columnas con encabezados 'fecha' y 'tasa' (case-insensitive).
    Si no se encuentra sheet_name se usa la primera hoja.
    """
    from openpyxl import load_workbook  # type: ignore[import-not-found]

    insertados = 0
    actualizados = 0
    errores: List[str] = []

    from io import BytesIO
    try:
        wb = load_workbook(filename=BytesIO(file_bytes))
    except Exception as e:  # noqa: BLE001
        return 0, 0, [f"Error leyendo XLSX: {e}"]

    ws = (
        wb[sheet_name]
        if sheet_name and sheet_name in wb.sheetnames
        else wb.active
    )
    if ws is None:
        return 0, 0, ["No se encontró una hoja activa en el archivo XLSX"]

    # Detectar encabezados
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=5):  # primeras filas
        values = [
            str(c.value).strip().lower() if c.value is not None else ""
            for c in row
        ]
        if "fecha" in values and "tasa" in values:
            header_row = row[0].row
            break
    if header_row is None:
        return 0, 0, [
            "No se encontró encabezado con columnas 'fecha' y 'tasa'"
        ]

    # Mapear índices
    headers = [
        str(c.value).strip().lower() if c.value is not None else ""
        for c in ws[header_row]
    ]
    try:
        idx_fecha = headers.index("fecha")
        idx_tasa = headers.index("tasa")
    except ValueError:
        return 0, 0, ["Encabezados inválidos"]

    for row in ws.iter_rows(min_row=header_row + 1):
        fecha_cell = row[idx_fecha].value
        tasa_cell = row[idx_tasa].value
        if fecha_cell in (None, "") or tasa_cell in (None, ""):
            continue
        try:
            # Convertir fecha si es datetime/date
            if isinstance(fecha_cell, datetime):
                fecha_dt_obj = fecha_cell.date()
            elif isinstance(fecha_cell, date):
                fecha_dt_obj = fecha_cell
            else:
                fecha_dt_obj = date.fromisoformat(str(fecha_cell).strip())

            if not isinstance(fecha_dt_obj, date):
                raise ValueError("Fecha inválida")

            tasa_val = float(str(tasa_cell).strip())
            creado, _ = upsert_tipo_cambio(
                conn,
                TipoCambioCreate(
                    fecha=fecha_dt_obj,
                    moneda=moneda,
                    tipo=tipo,
                    tasa=tasa_val,
                    origen=origen,
                    notas=None,
                ),
            )
            if creado:
                insertados += 1
            else:
                actualizados += 1
        except Exception as e:  # noqa: BLE001
            errores.append(f"Fila {row[0].row}: {e}")

    return insertados, actualizados, errores


def obtener_tasa_cercana(
    conn: SQLConn,
    moneda: str,
    fecha: date,
    tipo: str = "PROMEDIO",
) -> Optional[dict]:
    """Obtiene la tasa exacta o la más cercana a la fecha indicada.

    - Prioriza coincidencia exacta.
    - Si no existe, busca la última anterior.
    - Finalmente, intenta con la siguiente posterior.
    Devuelve None si no hay datos para la moneda/tipo.
    """

    params = {"moneda": moneda, "tipo": tipo, "fecha": fecha}
    base_sql = (
        "SELECT fecha, moneda, tipo, tasa, origen, notas "
        "FROM tipo_cambio_hist WHERE moneda=:moneda AND tipo=:tipo"
    )

    exact = conn.execute(
        text(f"{base_sql} AND fecha=:fecha ORDER BY fecha DESC LIMIT 1"),
        params,
    ).mappings().first()
    if exact:
        row = dict(exact)
        row["tasa"] = float(row["tasa"])
        row["fecha"] = row["fecha"]
        row["es_estimativa"] = False
        row["origen_busqueda"] = "exacta"
        return row

    prev = conn.execute(
        text(
            f"{base_sql} AND fecha < :fecha ORDER BY fecha DESC LIMIT 1"
        ),
        params,
    ).mappings().first()
    if prev:
        row = dict(prev)
        row["tasa"] = float(row["tasa"])
        row["es_estimativa"] = True
        row["origen_busqueda"] = "anterior"
        return row

    nxt = conn.execute(
        text(
            f"{base_sql} AND fecha > :fecha ORDER BY fecha ASC LIMIT 1"
        ),
        params,
    ).mappings().first()
    if nxt:
        row = dict(nxt)
        row["tasa"] = float(row["tasa"])
        row["es_estimativa"] = True
        row["origen_busqueda"] = "posterior"
        return row

    return None


def obtener_tasa_cercana_flexible(
    conn: SQLConn,
    moneda: str,
    fecha: date,
    tipos_prioridad: Sequence[str] | None = None,
) -> Optional[dict]:
    """Intenta obtener la tasa más cercana probando varios tipos (ordenados)."""
    tipos = tipos_prioridad or ("PROMEDIO", "VENTA", "COMPRA")
    for tipo in tipos:
        tasa = obtener_tasa_cercana(conn, moneda, fecha, tipo)
        if tasa:
            tasa["tipo_sugerido"] = tipo
            return tasa
    return None
