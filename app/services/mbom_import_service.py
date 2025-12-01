"""Servicios para importar estructuras MBOM desde planillas Flexxus."""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from typing import Any, Dict, List, Sequence, TypedDict

from fastapi import HTTPException, UploadFile
from sqlalchemy import text
from sqlalchemy.orm import Session

from . import mbom_service
from .producto_service import crear_producto, get_producto
from .unidad_service import listar_unidades


class ProductoLite(TypedDict):
    id: int
    codigo: str
    unidad_medida_id: int


@dataclass(slots=True)
class FlexxusRow:
    """Fila normalizada del archivo Flexxus."""

    line_number: int
    codigo: str
    descripcion: str
    cantidad: float | None
    nivel: int

    @property
    def normalized_codigo(self) -> str:
        return self.codigo.upper()


def importar_mbom_desde_flexxus(
    db: Session, producto_padre_id: int, archivo: UploadFile
) -> Dict[str, object]:
    """Importa componentes MBOM desde archivo CSV/XLSX de Flexxus."""
    if not archivo.filename:
        raise HTTPException(status_code=400, detail="Archivo requerido")

    prod_padre = get_producto(db, producto_padre_id)
    if not prod_padre:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    rows = _parse_upload(archivo)
    if not rows:
        raise HTTPException(status_code=400, detail="Archivo vacío")

    normalizados = _normalize_rows(rows)
    _validar_raiz(normalizados, prod_padre["codigo"])

    # Construir jerarquía: mapeo de código_padre -> lista de hijos
    jerarquia = _construir_jerarquia(normalizados)

    # Obtener o crear todos los productos necesarios (WIP intermedios + MP)
    unidades = listar_unidades(db)
    if not unidades:
        raise HTTPException(
            status_code=500,
            detail="No hay unidades de medida configuradas",
        )
    um_default_id = unidades[0]["id"]

    try:
        # Procesar jerarquía recursivamente desde el producto padre
        # Los hijos de nivel 1 se mapean con "__ROOT__" en construir_jerarquia
        hijos_raiz = jerarquia.get("__ROOT__", [])
        _procesar_nivel(
            db=db,
            codigo_padre=prod_padre["codigo"].upper(),
            hijos=hijos_raiz,
            jerarquia=jerarquia,
            um_default_id=um_default_id,
        )
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:  # pragma: no cover - safety net
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Error importando MBOM: {exc}",
        ) from exc

    # Retornar estructura del producto padre
    cabecera = mbom_service.obtener_o_crear_borrador(db, producto_padre_id)
    mbom_id = int(cabecera["id"])  # type: ignore[index]
    lineas = mbom_service.listar_lineas(db, mbom_id)
    cabecera_final = mbom_service.get_cabecera_por_id(db, mbom_id)
    return {"cabecera": cabecera_final, "lineas": lineas}


def _parse_upload(archivo: UploadFile) -> List[Dict[str, object]]:
    content = archivo.file.read()
    if not content:
        return []
    filename = (archivo.filename or "").lower()
    if filename.endswith(".xlsx"):
        return _parse_xlsx(content)
    return _parse_csv(content)


def _parse_csv(content: bytes) -> List[Dict[str, object]]:
    try:
        decoded = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            decoded = content.decode("latin-1")
        except UnicodeDecodeError as exc:
            raise HTTPException(
                status_code=400,
                detail=f"No se pudo decodificar el CSV: {exc}",
            ) from exc

    stream = io.StringIO(decoded)
    sample = stream.read(2048)
    stream.seek(0)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    reader = csv.DictReader(stream, dialect=dialect)
    rows: List[Dict[str, object]] = []
    for row in reader:
        cleaned: Dict[str, object] = {
            (k.strip() if isinstance(k, str) else k): (
                v.strip() if isinstance(v, str) else v
            )
            for k, v in row.items()
        }
        rows.append(cleaned)
    return rows


def _parse_xlsx(content: bytes) -> List[Dict[str, object]]:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # pragma: no cover - fallback de dependencia
        raise HTTPException(
            status_code=400,
            detail="Soporte XLSX no disponible en el servidor",
        ) from exc

    wb = openpyxl.load_workbook(
        io.BytesIO(content), read_only=True, data_only=True
    )
    ws: Any = wb.active
    rows: List[Dict[str, object]] = []
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [
        (str(cell.value).strip() if cell.value is not None else f"col_{idx}")
        for idx, cell in enumerate(header_row)
    ]
    for excel_row in ws.iter_rows(min_row=2):
        record: Dict[str, object] = {}
        for idx, cell in enumerate(excel_row):
            key = headers[idx] if idx < len(headers) else f"col_{idx}"
            value = cell.value
            record[key] = str(value).strip() if value is not None else ""
        rows.append(record)
    return rows


def _normalize_rows(rows: Sequence[Dict[str, object]]) -> List[FlexxusRow]:
    normalizados: List[FlexxusRow] = []
    for idx, row in enumerate(rows, start=2):
        lower_map = {
            (k.lower() if isinstance(k, str) else k): (
                str(v).strip() if v is not None else ""
            )
            for k, v in row.items()
        }
        codigo = (
            lower_map.get("codart")
            or lower_map.get("cod_art")
            or lower_map.get("codigo")
            or ""
        )
        descripcion = lower_map.get("descripcion") or ""
        nivel_raw = lower_map.get("nivel") or "0"
        try:
            nivel = int(float(nivel_raw.replace(",", "."))) if nivel_raw else 0
        except ValueError as exc:
            raise HTTPException(
                status_code=400,
                detail=f"Fila {idx}: Nivel inválido ({nivel_raw})",
            ) from exc
        cantidad_raw = lower_map.get("cantidad") or lower_map.get("cant") or ""
        cantidad = None
        if cantidad_raw:
            try:
                cantidad = float(
                    cantidad_raw.replace(" ", "").replace(",", ".")
                )
            except ValueError as exc:
                raise HTTPException(
                    status_code=400,
                    detail=f"Fila {idx}: Cantidad inválida ({cantidad_raw})",
                ) from exc
        normalizados.append(
            FlexxusRow(
                line_number=idx,
                codigo=codigo.strip(),
                descripcion=descripcion.strip(),
                cantidad=cantidad,
                nivel=nivel,
            )
        )
    return normalizados


def _validar_raiz(rows: List[FlexxusRow], codigo_padre: str) -> None:
    """Valida que la fila raíz coincida con el producto padre."""
    codigo_padre_upper = codigo_padre.upper()
    root = next((r for r in rows if r.nivel == 0 and r.codigo), None)
    if root and root.codigo.upper() != codigo_padre_upper:
        raise HTTPException(
            status_code=400,
            detail=(
                "El código raíz del archivo ("
                f"{root.codigo}) no coincide con el producto "
                f"seleccionado ({codigo_padre})"
            ),
        )


def _construir_jerarquia(
    rows: List[FlexxusRow],
) -> Dict[str, List[FlexxusRow]]:
    """Construye mapa código_padre -> [hijos directos]."""
    jerarquia: Dict[str, List[FlexxusRow]] = {}
    stack: List[FlexxusRow] = []  # Pila de padres por nivel

    for row in rows:
        if row.nivel == 0:
            continue  # Saltar raíz
        if row.codigo.startswith("PROCESO"):
            continue  # Ignorar procesos por ahora

        # Ajustar stack al nivel actual
        while len(stack) >= row.nivel:
            stack.pop()

        # Determinar padre: stack[-1] si existe, sino es hijo directo de raíz
        if stack:
            padre_codigo = stack[-1].normalized_codigo
        else:
            # Nivel 1: hijo directo de la raíz (se maneja en _procesar_nivel)
            padre_codigo = "__ROOT__"

        if padre_codigo not in jerarquia:
            jerarquia[padre_codigo] = []
        jerarquia[padre_codigo].append(row)

        # Agregar al stack si es un producto con subestructura potencial
        stack.append(row)

    return jerarquia


def _procesar_nivel(
    db: Session,
    codigo_padre: str,
    hijos: List[FlexxusRow],
    jerarquia: Dict[str, List[FlexxusRow]],
    um_default_id: int,
) -> None:
    """Procesa un nivel de la jerarquía, creando productos WIP y sus MBOMs."""
    # Obtener o crear producto padre
    producto_map = _fetch_productos_por_codigo(db, [codigo_padre])
    if codigo_padre not in producto_map:
        raise HTTPException(
            status_code=404,
            detail=f"Producto padre {codigo_padre} no encontrado",
        )

    prod_padre_id = producto_map[codigo_padre]["id"]

    # Crear/obtener MBOM BORRADOR para este producto
    cabecera = mbom_service.obtener_o_crear_borrador(db, prod_padre_id)
    mbom_id = int(cabecera["id"])  # type: ignore[index]

    # Limpiar detalle existente
    db.execute(
        text("DELETE FROM mbom_detalle WHERE mbom_id=:id"),
        {"id": mbom_id},
    )

    # Insertar componentes directos
    renglon = 10
    for hijo in hijos:
        if hijo.cantidad is None or hijo.cantidad <= 0:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Fila {hijo.line_number}: cantidad inválida "
                    f"para {hijo.codigo}"
                ),
            )

        # Obtener o crear producto hijo
        prod_hijo_map = _fetch_productos_por_codigo(
            db, [hijo.normalized_codigo]
        )
        if hijo.normalized_codigo not in prod_hijo_map:
            # Determinar tipo: WIP si tiene hijos, MP si no
            tiene_hijos = hijo.normalized_codigo in jerarquia
            tipo_prod = "WIP" if tiene_hijos else "MP"

            nuevo_prod = crear_producto(
                db=db,
                codigo=hijo.normalized_codigo,
                nombre=hijo.descripcion or hijo.normalized_codigo,
                tipo_producto=tipo_prod,
                unidad_medida_id=um_default_id,
                activo=True,
            )
            prod_hijo = ProductoLite(
                id=nuevo_prod["id"],
                codigo=nuevo_prod["codigo"],
                unidad_medida_id=nuevo_prod["unidad_medida_id"],
            )
        else:
            prod_hijo = prod_hijo_map[hijo.normalized_codigo]

        # Insertar línea en MBOM
        mbom_service.upsert_linea(
            db=db,
            mbom_id=mbom_id,
            renglon=renglon,
            componente_producto_id=prod_hijo["id"],
            cantidad=hijo.cantidad,
            unidad_medida_id=prod_hijo["unidad_medida_id"],
            notas=hijo.descripcion or None,
        )
        renglon += 10

        # Procesar recursivamente si tiene hijos
        if hijo.normalized_codigo in jerarquia:
            _procesar_nivel(
                db=db,
                codigo_padre=hijo.normalized_codigo,
                hijos=jerarquia[hijo.normalized_codigo],
                jerarquia=jerarquia,
                um_default_id=um_default_id,
            )


def _fetch_productos_por_codigo(
    db: Session, codigos_upper: Sequence[str]
) -> Dict[str, ProductoLite]:
    if not codigos_upper:
        return {}
    placeholders = ", ".join(f":c{i}" for i in range(len(codigos_upper)))
    params = {f"c{i}": codigo for i, codigo in enumerate(codigos_upper)}
    sql = text(
        "SELECT id, codigo, unidad_medida_id FROM producto "
        f"WHERE UPPER(codigo) IN ({placeholders})"
    )
    rows = db.execute(sql, params).fetchall()
    return {
        str(row.codigo).upper(): ProductoLite(
            id=int(row.id),
            codigo=str(row.codigo),
            unidad_medida_id=int(row.unidad_medida_id),
        )
        for row in rows
    }
