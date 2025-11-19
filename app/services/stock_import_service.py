from __future__ import annotations
from typing import List, Dict, Any, Optional
import csv
import io
from datetime import datetime

from fastapi import HTTPException, UploadFile
from sqlalchemy import text
from sqlalchemy.orm import Session

from ..schemas.stock import StockItemOut, StockImportResult


REQUIRED_STOCK_HEADERS = {"producto_codigo", "stock_disponible"}


def _get_producto_id(db: Session, codigo: str) -> Optional[int]:
    q = text("SELECT id FROM producto WHERE codigo = :codigo AND activo = 1")
    res = db.execute(q, {"codigo": codigo}).first()
    return int(res[0]) if res else None


def _parse_csv(content: bytes) -> List[Dict[str, Any]]:
    text_stream = io.StringIO(content.decode("utf-8-sig"))
    # Intentar detectar delimitador ; o ,
    sample = text_stream.read(2048)
    text_stream.seek(0)
    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    reader = csv.DictReader(text_stream, dialect=dialect)
    rows = [dict((k.strip(), v.strip() if isinstance(v, str) else v) for k, v in row.items()) for row in reader]
    return rows


def importar_stock_csv_o_excel(
    db: Session, anio: int, mes: int, archivo: UploadFile, fecha_corte: str
) -> StockImportResult:
    try:
        _ = datetime.strptime(fecha_corte, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="fecha_corte inválida (YYYY-MM-DD)")

    filename = archivo.filename or ""
    content = archivo.file.read()

    rows: List[Dict[str, Any]]
    if filename.lower().endswith(".csv") or filename == "":
        rows = _parse_csv(content)
    elif filename.lower().endswith(".xlsx"):
        try:
            import openpyxl  # type: ignore
        except Exception:
            raise HTTPException(status_code=400, detail="Soporte Excel no disponible; enviar CSV")
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for r in ws.iter_rows(min_row=2):
            row = {}
            for idx, cell in enumerate(r):
                key = headers[idx] if idx < len(headers) else f"col_{idx}"
                row[key] = str(cell.value).strip() if cell.value is not None else None
            rows.append(row)
    else:
        raise HTTPException(status_code=400, detail="Extensión no soportada (usar .csv o .xlsx)")

    # Validar encabezados
    if not rows:
        return StockImportResult(insertados=0, actualizados=0, rechazados=0, errores=["Archivo vacío"]) 
    headers_present = {k for k in rows[0].keys()}
    if not REQUIRED_STOCK_HEADERS.issubset({h.lower() for h in headers_present}):
        return StockImportResult(
            insertados=0,
            actualizados=0,
            rechazados=0,
            errores=[
                "Encabezados requeridos: producto_codigo, stock_disponible",
            ],
        )

    insertados = 0
    actualizados = 0
    rechazados = 0
    errores: List[str] = []

    for idx, row in enumerate(rows, start=2):  # 2 considerando encabezado
        codigo = (row.get("producto_codigo") or row.get("PRODUCTO_CODIGO") or "").strip()
        stock_raw = row.get("stock_disponible") or row.get("STOCK_DISPONIBLE")
        if not codigo:
            rechazados += 1
            errores.append(f"Fila {idx}: producto_codigo vacío")
            continue
        try:
            # Normalizar separador decimal
            stock = float(str(stock_raw).replace(",", "."))
            if stock < 0:
                raise ValueError()
        except Exception:
            rechazados += 1
            errores.append(f"Fila {idx}: stock_disponible inválido")
            continue

        prod_id = _get_producto_id(db, codigo)
        if not prod_id:
            rechazados += 1
            errores.append(f"Fila {idx}: producto no encontrado ({codigo})")
            continue

        # Upsert por periodo y producto
        upd = text(
            """
            UPDATE stock_disponible_mes
            SET stock_disponible=:stk, fecha_corte=:fc, origen='ERP_FLEXXUS'
            WHERE anio=:a AND mes=:m AND producto_id=:pid
            """
        )
        r = db.execute(
            upd, {"stk": stock, "fc": fecha_corte, "a": anio, "m": mes, "pid": prod_id}
        )
        if r.rowcount and r.rowcount > 0:
            actualizados += 1
        else:
            ins = text(
                """
                INSERT INTO stock_disponible_mes (anio, mes, producto_id, stock_disponible, fecha_corte, origen)
                VALUES (:a, :m, :pid, :stk, :fc, 'ERP_FLEXXUS')
                """
            )
            db.execute(ins, {"a": anio, "m": mes, "pid": prod_id, "stk": stock, "fc": fecha_corte})
            insertados += 1

    return StockImportResult(
        insertados=insertados, actualizados=actualizados, rechazados=rechazados, errores=errores
    )


def listar_stock_periodo(db: Session, anio: int, mes: int, q: Optional[str]) -> List[StockItemOut]:
    base_sql = (
        """
        SELECT s.id, s.anio, s.mes, p.codigo AS producto_codigo, s.stock_disponible, s.fecha_corte, s.origen
        FROM stock_disponible_mes s
        JOIN producto p ON p.id = s.producto_id
        WHERE s.anio = :a AND s.mes = :m
        """
    )
    params: Dict[str, Any] = {"a": anio, "m": mes}
    if q:
        base_sql += " AND (p.codigo LIKE :q OR p.nombre LIKE :q)"
        params["q"] = f"%{q}%"
    base_sql += " ORDER BY p.codigo"
    rows = db.execute(text(base_sql), params).mappings().all()
    return [
        StockItemOut(
            id=r["id"],
            anio=r["anio"],
            mes=r["mes"],
            producto_codigo=r["producto_codigo"],
            stock_disponible=r["stock_disponible"],
            fecha_corte=str(r["fecha_corte"]),
            origen=r["origen"],
        )
        for r in rows
    ]


def resumen_stock_periodo(db: Session, anio: int, mes: int) -> Dict[str, Any]:
    q = text(
        "SELECT COUNT(*) AS items, COALESCE(SUM(stock_disponible),0) AS total FROM stock_disponible_mes WHERE anio=:a AND mes=:m"
    )
    r = db.execute(q, {"a": anio, "m": mes}).first()
    return {"items": int(r[0]) if r else 0, "total_stock": float(r[1]) if r else 0.0}
