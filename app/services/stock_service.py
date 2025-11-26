import csv
from openpyxl import load_workbook, Workbook
from sqlalchemy.orm import Session
from app.schemas.stock import StockMensualImport, StockMensualOut
from app.models import StockDisponibleMes  # Asume modelo SQLAlchemy
from datetime import date
from decimal import Decimal

# Función para importar CSV

def bulk_import_csv(file_path: str, db: Session):
    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        items = []
        for row in reader:
            item = StockDisponibleMes(
                periodo=row['periodo'],
                codigo_producto=row['codigo_producto'],
                cantidad=float(row['cantidad']),
                unidad_medida=row['unidad_medida'],
                fecha_stock=row.get('fecha_stock')
            )
            items.append(item)
        db.bulk_save_objects(items)
        db.commit()
    return len(items)

# Función para importar XLSX

def bulk_import_xlsx(file_path: str, db: Session):
    wb = load_workbook(file_path)
    ws = wb.active
    items = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        periodo, codigo_producto, cantidad, unidad_medida, fecha_stock = row
        item = StockDisponibleMes(
            periodo=periodo,
            codigo_producto=codigo_producto,
            cantidad=float(cantidad),
            unidad_medida=unidad_medida,
            fecha_stock=fecha_stock
        )
        items.append(item)
    db.bulk_save_objects(items)
    db.commit()
    return len(items)

# Generar plantilla XLSX

def generar_template_xlsx(file_path: str):
    wb = Workbook()
    ws = wb.active
    ws.append(['periodo', 'codigo_producto', 'cantidad', 'unidad_medida', 'fecha_stock'])
    wb.save(file_path)

# Listar stock mensual

def listar_stock_mensual(db: Session):
    registros = db.query(StockDisponibleMes).all()
    resultado = []
    for r in registros:
        resultado.append({
            'id': r.id,
            'periodo': r.periodo,
            'codigo_producto': r.codigo_producto,
            'cantidad': float(r.cantidad) if isinstance(r.cantidad, Decimal) else r.cantidad,
            'unidad_medida': r.unidad_medida,
            'fecha_stock': r.fecha_stock.isoformat() if isinstance(r.fecha_stock, date) else r.fecha_stock
        })
    return resultado
