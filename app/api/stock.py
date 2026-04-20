from typing import List
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from ..db import get_db
from ..schemas.stock import StockItemOut, StockImportResult
from ..services.stock_import_service import (
    importar_stock_csv_o_excel,
    listar_stock_periodo,
    resumen_stock_periodo,
)
from .deps_auth import require_permission

router = APIRouter()


@router.get("/template-csv")
def descargar_template_csv(
    _current_user: dict = Depends(require_permission("stock", False)),
):
    contenido = (
        "producto_codigo,stock_disponible\n"
        "MP-0001,120\n"
        "MP-0002,85\n"
    )
    return StreamingResponse(
        io.BytesIO(contenido.encode("utf-8")),
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=template_stock_mensual.csv",
        },
    )


# Endpoint para descargar plantilla XLSX de stock mensual
@router.get("/template-xlsx")
def descargar_template_xlsx(
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("stock", False)),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.worksheet import Worksheet

    wb = Workbook()

    # ── Hoja 1: Instrucciones ────────────────────────────────────────────────
    ws_inst: Worksheet = wb.active  # type: ignore[assignment]
    ws_inst.title = "Instrucciones"
    ws_inst.sheet_view.showGridLines = False

    instrucciones = [
        ("PLANTILLA DE IMPORTACIÓN DE STOCK", "title"),
        ("", None),
        ("PASOS PARA IMPORTAR STOCK:", "section"),
        ("", None),
        ("1.  Ir a la hoja  'Carga'.", "normal"),
        ("2.  Pegar el listado exportado del ERP con las columnas:", "normal"),
        ("       •  codigo       (código del artículo en el sistema)", "indent"),
        ("       •  descripcion  (nombre, solo referencia visual)", "indent"),
        ("       •  stock        (cantidad disponible)", "indent"),
        ("", None),
        ("3.  Guardar el archivo.", "normal"),
        ("4.  Desde la UI: Stock → Importar → seleccionar este archivo.", "normal"),
        ("", None),
        ("NOTAS:", "section"),
        ("  •  El sistema reconoce la hoja 'Carga' automáticamente.", "normal"),
        ("  •  Solo se importan artículos cuyo código exista en el sistema.", "normal"),
        ("  •  La hoja 'Stock' lista las materias primas conocidas (solo consulta).", "normal"),
        ("  •  Columnas aceptadas en 'Carga': codigo / producto_codigo / code", "normal"),
        ("  •  Columnas aceptadas para stock: stock / stock_disponible / cantidad", "normal"),
    ]
    style_map = {
        "title":   Font(bold=True, size=14, color="0B5CAB"),
        "section": Font(bold=True, size=11, color="1E6B2F"),
        "normal":  Font(size=10),
        "indent":  Font(size=10, color="444444"),
    }
    for row_i, (txt, style) in enumerate(instrucciones, start=1):
        cell = ws_inst.cell(row=row_i, column=1, value=txt)
        if style:
            cell.font = style_map[style]
    ws_inst.column_dimensions["A"].width = 72

    # ── Hoja 2: Carga (input del usuario) ───────────────────────────────────
    ws_carga = wb.create_sheet("Carga")
    carga_fill = PatternFill("solid", fgColor="1E6B2F")
    carga_font = Font(bold=True, color="FFFFFF")
    carga_headers = ["codigo", "descripcion", "stock"]
    for col, title in enumerate(carga_headers, start=1):
        cell = ws_carga.cell(row=1, column=col, value=title)
        cell.fill = carga_fill
        cell.font = carga_font
        cell.alignment = Alignment(horizontal="center")
    ws_carga.column_dimensions["A"].width = 22
    ws_carga.column_dimensions["B"].width = 45
    ws_carga.column_dimensions["C"].width = 18
    ws_carga.freeze_panes = "A2"

    # ── Hoja 3: Stock (referencia — materias primas del sistema) ────────────
    ws_stock = wb.create_sheet("Stock")
    stock_fill = PatternFill("solid", fgColor="0B5CAB")
    stock_font = Font(bold=True, color="FFFFFF")
    stock_headers = ["producto_codigo", "nombre_articulo", "stock_disponible"]
    for col, title in enumerate(stock_headers, start=1):
        cell = ws_stock.cell(row=1, column=col, value=title)
        cell.fill = stock_fill
        cell.font = stock_font
        cell.alignment = Alignment(horizontal="center")

    ref_fill = PatternFill("solid", fgColor="F4F6F8")
    ref_font = Font(italic=True, color="555555")
    mp_rows = db.execute(
        text(
            "SELECT codigo, nombre FROM producto "
            "WHERE tipo_producto = 'MP' AND activo = 1 "
            "ORDER BY codigo"
        )
    ).fetchall()
    for row_idx, (codigo, nombre) in enumerate(mp_rows, start=2):
        ws_stock.cell(row=row_idx, column=1, value=codigo)
        ref_cell = ws_stock.cell(row=row_idx, column=2, value=nombre)
        ref_cell.fill = ref_fill
        ref_cell.font = ref_font
        ws_stock.cell(row=row_idx, column=3, value=0)
    ws_stock.column_dimensions["A"].width = 22
    ws_stock.column_dimensions["B"].width = 40
    ws_stock.column_dimensions["C"].width = 20
    ws_stock.freeze_panes = "A2"

    # Abrir en hoja Carga por defecto
    wb.active = ws_carga

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_stock_mensual.xlsx"},
    )


@router.post(
    "/{anio}/{mes}/import",
    response_model=StockImportResult,
    dependencies=[Depends(require_permission("stock", True))],
)
def importar_stock(
    anio: int,
    mes: int,
    archivo: UploadFile = File(...),
    fecha_corte: str = Form(..., description="Fecha de corte YYYY-MM-DD"),
    db: Session = Depends(get_db),
):
    if not 1 <= mes <= 12:
        raise HTTPException(
            status_code=400,
            detail="mes debe estar entre 1 y 12"
        )
    if not archivo.filename:
        raise HTTPException(status_code=400, detail="Archivo requerido")
    return importar_stock_csv_o_excel(db, anio, mes, archivo, fecha_corte)


@router.get(
    "/{anio}/{mes}",
    response_model=List[StockItemOut],
    dependencies=[Depends(require_permission("stock", False))],
)
def listar_stock(
    anio: int,
    mes: int,
    q: str | None = None,
    db: Session = Depends(get_db),
):
    if not (1 <= mes <= 12):
        raise HTTPException(
            status_code=400,
            detail="mes debe estar entre 1 y 12"
        )
    return listar_stock_periodo(db, anio, mes, q)


@router.get("/{anio}/{mes}/resumen")
def resumen_stock(
    anio: int,
    mes: int,
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("stock", False)),
):
    if not (1 <= mes <= 12):
        raise HTTPException(
            status_code=400,
            detail="mes debe estar entre 1 y 12"
        )
    return resumen_stock_periodo(db, anio, mes)
