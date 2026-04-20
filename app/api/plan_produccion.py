import csv
import io
from typing import List, Optional, cast

import openpyxl
from fastapi import (
    APIRouter,
    Body,
    Depends,
    File,
    HTTPException,
    Query,
    UploadFile,
)
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from app.api.deps import get_db
from app.api.deps_auth import require_permission
from app.services.plan_produccion_service import (
    calcular_asistente_oc,
    calcular_faltantes_y_capacidad,
    calcular_requerimientos_valorizados,
    guardar_deuda_clientes_periodo,
    guardar_bulk,
    crear_laf_solicitado_periodo,
    eliminar_laf_solicitado_periodo,
    guardar_stock_pt_periodo,
    importar_laf_solicitado_periodo,
    importar_desde_rows,
    listar_laf_solicitado_periodo,
    listar_corridas_asistente_oc,
    listar_periodos_cargados,
    listar_planes,
    obtener_ajustes_pt_periodo,
    registrar_corrida_asistente_oc,
    resumen_planes,
    resumen_rango_planes,
)

router = APIRouter(prefix="/plan-produccion-mensual", tags=["plan-produccion-mensual"])


def _parse_simple_csv(content: bytes) -> List[dict]:
    texto = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(texto))
    rows: List[dict] = []
    for r in reader:
        normalizado = {}
        for k, v in r.items():
            if not k:
                continue
            normalizado[str(k).strip().lower()] = v
        rows.append(normalizado)
    return rows


def _parse_simple_xlsx(content: bytes) -> List[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    if ws is None:
        return []

    headers: List[str] = []
    rows: List[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = ["" if v is None else str(v).strip() for v in row]
        if i == 1:
            headers = [v.lower() for v in values if str(v).strip()]
            continue

        if not headers:
            continue

        data_row: dict = {}
        for idx, h in enumerate(headers):
            data_row[h] = values[idx] if idx < len(values) else ""

        if any(str(v).strip() for v in data_row.values()):
            rows.append(data_row)
    return rows


def _parse_simple_upload(file_name: str, content: bytes) -> List[dict]:
    lower_name = (file_name or "").lower()
    if lower_name.endswith(".xlsx"):
        return _parse_simple_xlsx(content)
    return _parse_simple_csv(content)


@router.get("/", response_model=dict)
def listar(
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    mes: Optional[int] = None,
    anio: Optional[int] = None,
    producto_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", False)),
):
    items, total = listar_planes(
        db,
        limit=limit,
        offset=offset,
        mes=mes,
        anio=anio,
        producto_id=producto_id,
    )
    return {"items": items, "total": total}


@router.get("/periodos", response_model=dict)
def periodos_cargados(
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", False)),
):
    items = listar_periodos_cargados(db)
    return {"items": items, "total": len(items)}


@router.get("/resumen", response_model=dict)
def resumen(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", False)),
):
    items = resumen_planes(db, mes, anio)
    total_general = sum(i.get("cantidad", 0) for i in items)
    return {"items": items, "total_general": total_general}


@router.get("/requerimientos-valuados", response_model=dict)
def requerimientos_valuados(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
    persistir: bool = Query(False),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", False)),
):
    data = calcular_requerimientos_valorizados(db, mes, anio, persistir)
    return data


@router.get("/faltantes-capacidad", response_model=dict)
def faltantes_capacidad(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
    persistir_sugerencias: bool = Query(True),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", False)),
):
    permisos = _current_user.get("permissions", {}) if _current_user else {}
    permiso_plan = permisos.get("plan", (False, False))
    puede_escribir = bool(permiso_plan[1]) if isinstance(permiso_plan, (tuple, list)) else False

    if persistir_sugerencias and not puede_escribir:
        raise HTTPException(
            status_code=403,
            detail="No tenés permisos de escritura para persistir sugerencias.",
        )

    data = calcular_faltantes_y_capacidad(
        db,
        mes,
        anio,
        persistir_sugerencias,
    )
    return data


@router.post("/asistente-oc/calcular", response_model=dict)
def asistente_oc_calcular(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
    persistir_sugerencias: bool = Query(False),
    payload: dict = Body(default_factory=dict),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", False)),
):
    permisos = _current_user.get("permissions", {}) if _current_user else {}
    permiso_plan = permisos.get("plan", (False, False))
    puede_escribir = bool(permiso_plan[1]) if isinstance(permiso_plan, (tuple, list)) else False

    if persistir_sugerencias and not puede_escribir:
        raise HTTPException(
            status_code=403,
            detail="No tenés permisos de escritura para persistir sugerencias.",
        )

    ajustes_pt = payload.get("ajustes_pt") if isinstance(payload, dict) else None
    if ajustes_pt is not None and not isinstance(ajustes_pt, list):
        raise HTTPException(
            status_code=400,
            detail="El campo ajustes_pt debe ser una lista.",
        )

    data = calcular_asistente_oc(
        db,
        mes,
        anio,
        ajustes_pt=ajustes_pt or [],
        persistir_sugerencias=persistir_sugerencias,
    )

    usuario_id = int(_current_user.get("id") or 0)
    usuario_email = str(_current_user.get("email") or "")
    usuario_nombre = _current_user.get("nombre")
    if usuario_id > 0 and usuario_email:
        try:
            corrida_id = registrar_corrida_asistente_oc(
                db,
                mes,
                anio,
                usuario_id=usuario_id,
                usuario_email=usuario_email,
                usuario_nombre=str(usuario_nombre or "") or None,
                persistio_sugerencias=persistir_sugerencias,
            )
            data["corrida"] = {
                "id": corrida_id,
                "usuario_id": usuario_id,
                "usuario_email": usuario_email,
                "usuario_nombre": str(usuario_nombre or ""),
            }
        except SQLAlchemyError:
            # El cálculo no debe fallar si no se puede auditar la corrida.
            pass

    return data


@router.get("/asistente-oc/ajustes", response_model=dict)
def asistente_oc_ajustes(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", False)),
):
    items = obtener_ajustes_pt_periodo(db, mes, anio)
    return {"items": items, "total": len(items)}


@router.get("/asistente-oc/laf-solicitado", response_model=dict)
def asistente_oc_laf_solicitado(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", False)),
):
    items = listar_laf_solicitado_periodo(db, mes, anio)
    total_solicitado = sum(float(i.get("cantidad_total") or 0.0) for i in items)
    return {
        "items": items,
        "total": len(items),
        "total_solicitado": total_solicitado,
    }


@router.post("/asistente-oc/laf-solicitado", response_model=dict)
def asistente_oc_laf_solicitado_crear(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
    payload: dict = Body(default_factory=dict),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", True)),
):
    try:
        item = crear_laf_solicitado_periodo(db, mes, anio, payload or {})
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {"item": item}


@router.post("/asistente-oc/import-laf-solicitado", response_model=dict)
async def asistente_oc_import_laf_solicitado(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", True)),
):
    contenido = await file.read()
    rows = _parse_simple_upload(file.filename or "", contenido)
    items: List[dict] = []
    for r in rows:
        items.append(
            {
                "codigo": r.get("codigo") or r.get("producto") or r.get("producto_codigo"),
                "proveedor_nombre": r.get("proveedor") or r.get("proveedor_nombre"),
                "cantidad_total": r.get("cantidad_total") or r.get("cantidad"),
                "cantidad_q1": r.get("cantidad_q1") or r.get("q1") or r.get("quincena_1"),
                "cantidad_q2": r.get("cantidad_q2") or r.get("q2") or r.get("quincena_2"),
                "fecha_pedido": r.get("fecha_pedido"),
                "fecha_entrega_estimada": r.get("fecha_entrega_estimada") or r.get("fecha_entrega"),
                "estado": r.get("estado"),
                "observaciones": r.get("observaciones"),
            }
        )

    data = importar_laf_solicitado_periodo(db, mes, anio, items)
    return data


@router.delete("/asistente-oc/laf-solicitado/{item_id}", response_model=dict)
def asistente_oc_laf_solicitado_eliminar(
    item_id: int,
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", True)),
):
    ok = eliminar_laf_solicitado_periodo(db, item_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    return {"deleted": True}


@router.get("/asistente-oc/corridas", response_model=dict)
def asistente_oc_corridas(
    mes: Optional[int] = Query(None, ge=1, le=12),
    anio: Optional[int] = Query(None, ge=2000, le=2100),
    limit: int = Query(20, ge=1, le=200),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", False)),
):
    try:
        items, total = listar_corridas_asistente_oc(
            db,
            mes=mes,
            anio=anio,
            limit=limit,
            offset=offset,
        )
    except SQLAlchemyError:
        items, total = [], 0
    return {
        "items": items,
        "total": total,
        "limit": limit,
        "offset": offset,
    }


@router.post("/asistente-oc/import-stock-pt", response_model=dict)
async def asistente_oc_import_stock_pt(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
    fecha_corte: Optional[str] = Query(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", True)),
):
    contenido = await file.read()
    rows = _parse_simple_upload(file.filename or "", contenido)
    items: List[dict] = []
    for r in rows:
        items.append(
            {
                "codigo": r.get("codigo") or r.get("producto") or r.get("producto_codigo"),
                "stock_pt": r.get("stock_pt") or r.get("stock") or r.get("cantidad"),
            }
        )

    data = guardar_stock_pt_periodo(db, mes, anio, fecha_corte, items)
    return data


@router.post("/asistente-oc/import-deuda-clientes", response_model=dict)
async def asistente_oc_import_deuda_clientes(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
    fecha_corte: Optional[str] = Query(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", True)),
):
    contenido = await file.read()
    rows = _parse_simple_upload(file.filename or "", contenido)
    items: List[dict] = []
    for r in rows:
        items.append(
            {
                "codigo": r.get("codigo") or r.get("producto") or r.get("producto_codigo"),
                "deuda_clientes": (
                    r.get("deuda_clientes")
                    or r.get("deuda")
                    or r.get("pendiente")
                    or r.get("cantidad")
                ),
            }
        )

    data = guardar_deuda_clientes_periodo(db, mes, anio, fecha_corte, items)
    return data


@router.get("/asistente-oc/plantilla-stock-pt.csv")
def asistente_oc_plantilla_stock_pt(
    _current_user: dict = Depends(require_permission("plan", False)),
):
    contenido = (
        "codigo,stock_pt\n"
        "PT-0001,15\n"
        "PT-0002,30\n"
    )
    return StreamingResponse(
        io.BytesIO(contenido.encode("utf-8")),
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=plantilla_stock_pt.csv",
        },
    )


@router.get("/asistente-oc/plantilla-stock-pt.xlsx")
def asistente_oc_plantilla_stock_pt_xlsx(
    _current_user: dict = Depends(require_permission("plan", False)),
):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Stock PT"
    ws.append(["codigo", "stock_pt"])
    ws.append(["PT-0001", 15])
    ws.append(["PT-0002", 30])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=plantilla_stock_pt.xlsx",
        },
    )


@router.get("/asistente-oc/plantilla-deuda-clientes.csv")
def asistente_oc_plantilla_deuda_clientes(
    _current_user: dict = Depends(require_permission("plan", False)),
):
    contenido = (
        "codigo,deuda_clientes\n"
        "PT-0001,20\n"
        "PT-0002,5\n"
    )
    return StreamingResponse(
        io.BytesIO(contenido.encode("utf-8")),
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=plantilla_deuda_clientes.csv",
        },
    )


@router.get("/asistente-oc/plantilla-deuda-clientes.xlsx")
def asistente_oc_plantilla_deuda_clientes_xlsx(
    _current_user: dict = Depends(require_permission("plan", False)),
):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Deuda Clientes"
    ws.append(["codigo", "deuda_clientes"])
    ws.append(["PT-0001", 20])
    ws.append(["PT-0002", 5])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=plantilla_deuda_clientes.xlsx",
        },
    )


@router.get("/asistente-oc/plantilla-laf-solicitado.csv")
def asistente_oc_plantilla_laf_solicitado_csv(
    _current_user: dict = Depends(require_permission("plan", False)),
):
    contenido = (
        "codigo,proveedor_nombre,cantidad_total,cantidad_q1,cantidad_q2,"
        "fecha_pedido,fecha_entrega_estimada,estado,observaciones\n"
        "CHAPA-LAF-01,Proveedor LAF,1200,600,600,2026-01-05,2026-04-08,PENDIENTE,Pedido trimestral\n"
    )
    return StreamingResponse(
        io.BytesIO(contenido.encode("utf-8")),
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=plantilla_laf_solicitado.csv",
        },
    )


@router.get("/asistente-oc/plantilla-laf-solicitado.xlsx")
def asistente_oc_plantilla_laf_solicitado_xlsx(
    _current_user: dict = Depends(require_permission("plan", False)),
):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Solicitado LAF"
    ws.append(
        [
            "codigo",
            "proveedor_nombre",
            "cantidad_total",
            "cantidad_q1",
            "cantidad_q2",
            "fecha_pedido",
            "fecha_entrega_estimada",
            "estado",
            "observaciones",
        ]
    )
    ws.append(
        [
            "CHAPA-LAF-01",
            "Proveedor LAF",
            1200,
            600,
            600,
            "2026-01-05",
            "2026-04-08",
            "PENDIENTE",
            "Pedido trimestral",
        ]
    )

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=plantilla_laf_solicitado.xlsx",
        },
    )


@router.get("/asistente-oc/export-propuesta.csv")
def asistente_oc_export_propuesta(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", False)),
):
    data = calcular_asistente_oc(
        db,
        mes,
        anio,
        ajustes_pt=[],
        persistir_sugerencias=False,
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "codigo",
        "detalle",
        "um",
        "cantidad_requerida",
        "stock_disponible",
        "faltante",
        "cobertura_pct",
    ])

    for it in data.get("faltantes", []):
        writer.writerow(
            [
                it.get("codigo"),
                it.get("nombre"),
                it.get("um_codigo"),
                it.get("cantidad_requerida"),
                it.get("stock_disponible"),
                it.get("faltante"),
                it.get("cobertura_pct"),
            ]
        )

    payload = output.getvalue().encode("utf-8")
    filename = f"propuesta_oc_{anio}_{mes:02d}.csv"
    return StreamingResponse(
        io.BytesIO(payload),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
        },
    )


@router.get("/asistente-oc/export-propuesta.xlsx")
def asistente_oc_export_propuesta_xlsx(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", False)),
):
    data = calcular_asistente_oc(
        db,
        mes,
        anio,
        ajustes_pt=[],
        persistir_sugerencias=False,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Propuesta OC"

    ws.append(
        [
            "Codigo",
            "Detalle",
            "UM",
            "Cantidad requerida",
            "Stock disponible",
            "Faltante",
            "Cobertura %",
        ]
    )

    for it in data.get("faltantes", []):
        ws.append(
            [
                it.get("codigo"),
                it.get("nombre"),
                it.get("um_codigo"),
                it.get("cantidad_requerida"),
                it.get("stock_disponible"),
                it.get("faltante"),
                it.get("cobertura_pct"),
            ]
        )

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"propuesta_oc_{anio}_{mes:02d}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
        },
    )


@router.get("/requerimientos-valuados.xlsx")
def requerimientos_valuados_xlsx(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
    persistir: bool = Query(False),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", False)),
):
    data = calcular_requerimientos_valorizados(db, mes, anio, persistir)

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Requerimientos"
    headers = [
        "Codigo",
        "Nombre",
        "UM",
        "Cantidad",
        "Precio unit USD",
        "Precio unit ARS",
        "Total USD",
        "Total ARS",
        "Fuente",
        "Moneda origen",
        "Fecha precio",
        "FX USDtoARS",
        "FX estimada",
    ]
    ws.append(headers)

    for it in data.get("items", []):
        ws.append(
            [
                it.get("codigo"),
                it.get("nombre"),
                it.get("um_codigo"),
                float(it.get("cantidad") or 0),
                it.get("precio_unit_usd"),
                it.get("precio_unit_ars"),
                it.get("total_usd"),
                it.get("total_ars"),
                it.get("fuente"),
                it.get("moneda_origen"),
                it.get("fecha_precio"),
                it.get("fx_tasa_usd_ars"),
                it.get("fx_es_estimativa"),
            ]
        )

    ws.append([])
    ws.append([
        "",
        "",
        "",
        "",
        "",
        "Totales:",
        data.get("total_usd"),
        data.get("total_ars"),
    ])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    fname = f"requerimientos_{anio}_{mes:02d}.xlsx"
    headers_resp = {
        "Content-Disposition": (
            f"attachment; filename={fname}; filename*=UTF-8''{fname}"
        ),
        "Content-Type": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    }
    return StreamingResponse(
        stream,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers=headers_resp,
    )


@router.get("/resumen-rango", response_model=dict)
def resumen_rango(
    desde_mes: int = Query(..., ge=1, le=12),
    desde_anio: int = Query(..., ge=2000, le=2100),
    hasta_mes: int = Query(..., ge=1, le=12),
    hasta_anio: int = Query(..., ge=2000, le=2100),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", False)),
):
    try:
        data = resumen_rango_planes(
            db,
            desde_mes,
            desde_anio,
            hasta_mes,
            hasta_anio,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return data


@router.post("/bulk", response_model=dict)
def guardar_en_lote(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
    items: List[dict] = Body(default_factory=list),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", True)),
):
    count = guardar_bulk(db, mes, anio, items)
    return {"procesados": count}


@router.post("/import", response_model=dict)
async def importar_archivo(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(require_permission("plan", True)),
):
    contenido = await file.read()
    nombre = (file.filename or "").lower()
    rows: List[dict] = []

    if nombre.endswith(".csv"):
        texto = contenido.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(texto))
        for r in reader:
            rows.append(
                {
                    "codigo": r.get("Codigo") or r.get("codigo"),
                    "mes": r.get("Mes") or r.get("mes"),
                    "anio": r.get("Año") or r.get("Anio") or r.get("anio"),
                    "cantidad": r.get("Cantidad") or r.get("cantidad"),
                }
            )
    else:
        wb = openpyxl.load_workbook(io.BytesIO(contenido))
        sheet = cast(Worksheet, wb.active)
        headers = [str(c.value).strip() if c.value else "" for c in next(sheet.rows)]
        idx = {h.lower(): i for i, h in enumerate(headers)}

        def tomar(row, key: str):
            pos = idx.get(key)
            if pos is None:
                return None
            return row[pos].value

        for fila in sheet.iter_rows(min_row=2):
            rows.append(
                {
                    "codigo": tomar(fila, "codigo"),
                    "mes": tomar(fila, "mes"),
                    "anio": tomar(fila, "año") or tomar(fila, "anio"),
                    "cantidad": tomar(fila, "cantidad"),
                }
            )

    procesadas = importar_desde_rows(db, rows)
    return {"procesadas": procesadas}


@router.get("/plantilla.csv")
def plantilla_csv(
    _current_user: dict = Depends(require_permission("plan", False)),
):
    return FileResponse(
        "import/plan_produccion_template.csv",
        media_type="text/csv",
        filename="plan_produccion_template.csv",
    )


@router.get("/plantilla.xlsx")
def plantilla_xlsx(
    _current_user: dict = Depends(require_permission("plan", False)),
):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Plan"
    headers = ["Codigo", "Nombre", "Mes", "Año", "Cantidad"]
    ws.append(headers)
    ws.append(["PT-0001", "Producto Terminado Ejemplo 1", 12, 2025, 100])
    ws.append(["PT-0002", "Producto Terminado Ejemplo 2", 12, 2025, 200])
    ws.append(["PT-0003", "Producto Terminado Ejemplo 3", 12, 2025, 0])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    headers_resp = {
        "Content-Disposition": (
            "attachment; filename=plan_produccion_template.xlsx; "
            "filename*=UTF-8''plan_produccion_template.xlsx"
        ),
        "Content-Type": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    }

    return StreamingResponse(
        stream,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers=headers_resp,
    )
