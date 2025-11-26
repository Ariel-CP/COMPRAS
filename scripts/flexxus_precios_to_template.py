"""Utility to convert Flexxus export files into the price-import template.

Usage:
    python scripts/flexxus_precios_to_template.py --input Precios.xlsx --output precios_limpios.xlsx

The script keeps only rows whose producto_codigo exists in the local DB and
produces a file compatible with /api/precios/import.
"""
from __future__ import annotations

import argparse
import csv
import io
import sys
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Mapping

from sqlalchemy import text

try:
    import openpyxl  # type: ignore
except ImportError as exc:  # pragma: no cover - optional dependency guard
    raise RuntimeError(
        "openpyxl is required. Install with `pip install openpyxl`."
    ) from exc

# Ensure repository root is on sys.path so we can import app.*
ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from app.db import SessionLocal  # type: ignore  # noqa: E402

TEMPLATE_COLUMNS = [
    "producto_codigo",
    "proveedor_codigo",
    "proveedor_nombre",
    "fecha_precio",
    "precio_unitario",
    "moneda",
    "origen",
    "referencia_doc",
    "notas",
]

COLUMN_ALIASES = {
    "producto_codigo": [
        "producto_codigo",
        "codigo_articulo",
        "cod_articulo",
        "articulo",
        "item",
        "codigo",
    ],
    "proveedor_codigo": [
        "proveedor_codigo",
        "codigo_proveedor",
        "cod_proveedor",
        "proveedor",
    ],
    "proveedor_nombre": [
        "proveedor_nombre",
        "nombre_proveedor",
        "prov_nombre",
    ],
    "fecha_precio": [
        "fecha_precio",
        "fecha_compra",
        "fecha",
        "fecha_oc",
    ],
    "precio_unitario": [
        "precio_unitario",
        "precio",
        "importe",
        "precio_compra",
    ],
    "moneda": ["moneda", "divisa"],
    "origen": ["origen", "fuente"],
    "referencia_doc": ["referencia", "referencia_doc", "oc", "pedido"],
    "notas": ["notas", "observaciones"],
    "tipo_cambio": ["tipo_cambio", "tc", "cotizacion"],
}

DEFAULT_ORIGEN = "ERP_FLEXXUS"
DEFAULT_MONEDA = "ARS"
FALLBACK_PROV_CODIGO = "PROV_GENERICO"
FALLBACK_PROV_NOMBRE = "Proveedor Genérico"

RowDict = Dict[str, Any]


def _lower_keys(row: Mapping[Any, Any]) -> RowDict:
    result: RowDict = {}
    for key, value in row.items():
        if key is None:
            continue
        trimmed_key = str(key).strip().lower()
        cleaned_value = value.strip() if isinstance(value, str) else value
        result[trimmed_key] = cleaned_value
    return result


def _decode_csv_bytes(raw: bytes) -> str:
    try:
        return raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        return raw.decode("latin-1")


def _read_csv(path: Path) -> List[RowDict]:
    content = path.read_bytes()
    if not content:
        raise ValueError("El archivo CSV está vacío")
    text_stream = io.StringIO(_decode_csv_bytes(content))
    sample = text_stream.read(2048)
    text_stream.seek(0)
    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    reader = csv.DictReader(text_stream, dialect=dialect)
    return [_lower_keys(row) for row in reader]


def _read_xlsx(path: Path) -> List[RowDict]:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("El archivo XLSX no tiene hojas activas")
    try:
        headers = [
            str(c.value).strip().lower() if c.value is not None else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]
    except StopIteration:
        return []
    rows: List[RowDict] = []
    for xl_row in ws.iter_rows(min_row=2, values_only=True):
        row_dict: RowDict = {}
        for idx, cell in enumerate(xl_row):
            key = headers[idx] if idx < len(headers) else f"col_{idx}"
            row_dict[key] = cell
        rows.append(_lower_keys(row_dict))
    return rows


def _resolve_value(row: RowDict, field: str) -> Optional[Any]:
    for alias in COLUMN_ALIASES.get(field, [field]):
        if alias in row and row[alias] not in (None, ""):
            return row[alias]
    return None


def _to_date(value: object) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, date):
        return value
    value_str = str(value).strip()
    if not value_str:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y%m%d"):
        try:
            return datetime.strptime(value_str, fmt).date()
        except ValueError:
            continue
    return None


def _to_float(value: object) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


@dataclass
class CleansedRow:
    producto_codigo: str
    proveedor_codigo: str
    proveedor_nombre: Optional[str]
    fecha_precio: date
    precio_unitario: float
    moneda: str
    origen: str
    referencia_doc: Optional[str]
    notas: Optional[str]

    def as_list(self) -> List[object]:
        return [
            self.producto_codigo,
            self.proveedor_codigo,
            self.proveedor_nombre or "",
            self.fecha_precio.isoformat(),
            f"{self.precio_unitario:.6f}",
            self.moneda,
            self.origen,
            self.referencia_doc or "",
            self.notas or "",
        ]


def load_product_codes() -> set[str]:
    with SessionLocal() as session:
        rows = session.execute(
            text("SELECT codigo FROM producto")
        )  # type: ignore[name-defined]
        return {row[0] for row in rows}


def cleanse_rows(
    rows: Iterable[RowDict],
    product_codes: set[str],
    default_proveedor_codigo: str,
    default_proveedor_nombre: str,
) -> tuple[List[CleansedRow], List[str]]:
    cleansed: List[CleansedRow] = []
    rejected: List[str] = []
    for idx, row in enumerate(rows, start=2):
        codigo = _resolve_value(row, "producto_codigo")
        if not codigo:
            rejected.append(f"Fila {idx}: producto_codigo vacío")
            continue
        codigo_str = str(codigo).strip()
        if codigo_str not in product_codes:
            rejected.append(
                f"Fila {idx}: producto {codigo_str} no existe en la base"
            )
            continue

        proveedor_codigo = _resolve_value(row, "proveedor_codigo")
        if not proveedor_codigo:
            proveedor_codigo = default_proveedor_codigo
            proveedor_nombre = (
                _resolve_value(row, "proveedor_nombre")
                or default_proveedor_nombre
            )
        else:
            proveedor_nombre = (
                _resolve_value(row, "proveedor_nombre")
                or default_proveedor_nombre
            )

        fecha_precio = _to_date(_resolve_value(row, "fecha_precio"))
        if not fecha_precio:
            rejected.append(f"Fila {idx}: fecha_precio inválida")
            continue

        precio_unitario = _to_float(_resolve_value(row, "precio_unitario"))
        if not precio_unitario or precio_unitario <= 0:
            rejected.append(f"Fila {idx}: precio_unitario inválido")
            continue

        moneda = str(_resolve_value(row, "moneda") or DEFAULT_MONEDA).upper()
        origen = str(_resolve_value(row, "origen") or DEFAULT_ORIGEN).upper()
        referencia = _resolve_value(row, "referencia_doc")
        notas_raw = _resolve_value(row, "notas")
        tipo_cambio = _resolve_value(row, "tipo_cambio")
        notas = str(notas_raw) if notas_raw else ""
        if tipo_cambio not in (None, ""):
            notas = (notas + " ").strip()
            notas = f"{notas}TC={tipo_cambio}".strip()

        cleansed.append(
            CleansedRow(
                producto_codigo=codigo_str,
                proveedor_codigo=str(proveedor_codigo).strip(),
                proveedor_nombre=(
                    str(proveedor_nombre).strip() if proveedor_nombre else None
                ),
                fecha_precio=fecha_precio,
                precio_unitario=precio_unitario,
                moneda=moneda,
                origen=origen,
                referencia_doc=str(referencia).strip() if referencia else None,
                notas=notas or None,
            )
        )
    return cleansed, rejected


def write_output(rows: List[CleansedRow], destination: Path) -> None:
    if destination.suffix.lower() == ".xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise ValueError("No se pudo crear la hoja activa en el XLSX")
        ws.append(TEMPLATE_COLUMNS)
        for row in rows:
            ws.append(row.as_list())
        wb.save(destination)
    else:
        with destination.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(TEMPLATE_COLUMNS)
            for row in rows:
                writer.writerow(row.as_list())


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Limpia un archivo de Flexxus y genera uno listo para importar"
        )
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Ruta al archivo CSV/XLSX exportado de Flexxus",
    )
    parser.add_argument(
        "--output",
        default="precios_importables.xlsx",
        help="Ruta del archivo de salida (CSV o XLSX)",
    )
    parser.add_argument(
        "--default-proveedor-codigo",
        default=FALLBACK_PROV_CODIGO,
        help="Código a usar cuando no venga proveedor_codigo",
    )
    parser.add_argument(
        "--default-proveedor-nombre",
        default=FALLBACK_PROV_NOMBRE,
        help="Nombre a usar cuando no venga proveedor_nombre",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    if not input_path.exists():
        raise SystemExit(f"No se encontró el archivo {input_path}")

    if input_path.suffix.lower() == ".csv":
        raw_rows = _read_csv(input_path)
    elif input_path.suffix.lower() in {".xlsx", ".xls"}:
        raw_rows = _read_xlsx(input_path)
    else:
        raise SystemExit("Formato no soportado. Use CSV o XLSX")

    product_codes = load_product_codes()
    cleansed, rejected = cleanse_rows(
        raw_rows,
        product_codes,
        args.default_proveedor_codigo,
        args.default_proveedor_nombre,
    )

    if not cleansed:
        raise SystemExit(
            "No hay filas válidas para exportar. Revise los errores: \n"
            + "\n".join(rejected[:20])
        )

    write_output(cleansed, output_path)

    print(f"Generado {output_path} con {len(cleansed)} filas válidas")
    if rejected:
        print(f"Se descartaron {len(rejected)} filas:")
        for err in rejected[:20]:
            print(f" - {err}")
        if len(rejected) > 20:
            print("   ... (verifique el archivo original para más detalles)")


if __name__ == "__main__":
    main()
